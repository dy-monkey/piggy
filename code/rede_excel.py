import pandas as pd
from pandas import to_datetime
from datetime import datetime as dt

from openpyxl.utils.dataframe import dataframe_to_rows
from  openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border,Side

# options.io.excel.xlsx.writer = 'openpyxl'


with pd.ExcelFile('./useful.xlsx') as  xlsx:
    df1 = pd.read_excel(xlsx, sheet_name=4)
    df2 = pd.read_excel(xlsx, sheet_name=5)
df2 = df2[['产品编码', '产品名称','规格']]
dict = {}

for index, row in df1.iterrows():
    row['日期'] = to_datetime(row['日期']).strftime('%Y-%m-%d')
    print(type(row['日期']))
    print(row['日期'])
    key = str(row['日期']) + row['收货人']
    if key in dict.keys():
        dict[key].append({'产品编码':row['产品编码'],'数量': row['数量'], '单价': row['商品销售价'], '金额': row['数量'] * row['商品销售价']})
    else:
        dict[key] = []
        dict[key].append({'产品编码':row['产品编码'],'数量': row['数量'], '单价': row['商品销售价'], '金额': row['数量'] * row['商品销售价']})

# for key,value in dict:
dataframe = pd.DataFrame(dict['2020-04-10顾静妹'])
result = pd.merge(dataframe,df2, on='产品编码', how='left')
result['备注'] = ""
result.index=result.index+1

result.index.name = '序号'

result = result[['序号', '产品编码', '产品名称', '规格型号', '单位', '数量', '单价', '金额', '备注']]



writer = pd.ExcelWriter('test1.xlsx', engine='openpyxl')
result.to_excel(writer, sheet_name='Sheet1')
wb = writer.book
# writer.book = book
ws = wb.worksheets[0]

font = Font(name='Calibri',size='20',bold=True)
align = Alignment(horizontal='center')
border = Border(left=Side(border_style='thin',color='000000'),
right=Side(border_style='thin',color='000000'),
top=Side(border_style='thin',color='000000'),
bottom=Side(border_style='thin',color='000000'))

# for i in range(4):
ws.insert_rows(0,4)
ws.merge_cells('A1:H1')
ws['A1'] = '杭州新颖饮业有限公司'
ws['A1'].font = font
ws['A1'].alignment = align
ws.merge_cells('A2:H2')
ws['A2'] = '出售单（客户联）'
ws['A2'].font = font
ws['A2'].alignment = align
ws.merge_cells('A3:H3')
ws['A3'] = '单据编号'
ws['A3'].alignment = Alignment(horizontal='right')
ws.merge_cells('A4:E4')
ws.merge_cells('F4:H4')
ws['A4'] = '购货单位：'
ws['F4'] = '日期'
ws['F4'].alignment = Alignment(horizontal='right')
num = ws.max_row

ws.merge_cells('A{}:B{}'.format(num+1,num+1))
ws.merge_cells('C{}:F{}'.format(num+1,num+1))
ws.merge_cells('G{}:H{}'.format(num+1,num+1))
ws.merge_cells('A{}:H{}'.format(num+2,num+2))
# print(ws.column_dimensions['A{}'.format(num+2)].width )
ws['A{}'.format(num+2)] = "经办人:沈兆君" + "批准人：" + "客户签收："
ws['A{}'.format(num+1)] = "合计（大写）："
ws['G{}'.format(num+1)] = "（小写）￥："

for row in ws.rows:
    for cell in row:
        cell.border =border

length = 0
for column in ws.columns:
    column = column[4:num]
    for x in column:
        x.alignment = align
    temp_num = max([len(str(x.value)) for x in column])
    if temp_num != 0:
        ws.column_dimensions[column[0].column_letter].width = 2.5 * temp_num
    length += temp_num * 2.5

ws['A{}'.format(num+2)] = "经办人:      沈兆君" + int(length/2) * " " + "批准人：" + int(length/2) * " " + "客户签收："
ws.column_dimensions['H'].width = 25
wb.save('test1.xlsx')



print(dict)




def write_excel(dataframe,key):
    pass

# def dict_to_dataframe()
# # def

# def append_two_row(obj,row_num):
